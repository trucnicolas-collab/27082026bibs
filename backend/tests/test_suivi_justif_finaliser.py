"""Tests for new suivi-terrain features: justification 5%, a_finaliser, extra_products, fixations_reel, Excel export (3 sheets)."""
import io
import os
import openpyxl
import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL").rstrip("/")
UPLOAD_ID = "fd15443f-6d2a-4cef-bd72-c56bb29e9c42"
UID = "1__A__R1"
NIGHT = 2
T = f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}"


@pytest.fixture(scope="module")
def auth():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login",
               json={"email": "admin@vusion.local", "password": "admin123"})
    assert r.status_code == 200, r.text
    return s


@pytest.fixture(scope="module", autouse=True)
def ensure_published(auth):
    auth.post(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/publish", json={"published": True})
    yield
    # Cleanup: reset allée to a_faire, no justif, no extras
    requests.patch(f"{T}/allee", json={
        "uid": UID, "status": "a_faire", "justification": "",
        "extra_products": [],
        "products": [{"designation": "ES 1.5 noir", "reel": None}],
    })
    requests.patch(f"{T}/allee-cam", json={"uid": UID, "fixations_reel": None})
    auth.post(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/publish", json={"published": True})


# ---------- State exposes new fields ----------
def test_state_exposes_new_fields():
    j = requests.get(T).json()
    assert j.get("allees")
    a = next(x for x in j["allees"] if x["uid"] == UID)
    for k in ("justification", "justif_products", "extra_products"):
        assert k in a
    # night has nb_a_finaliser
    n = next(n for n in j["nights"] if n["nuit"] == NIGHT)
    assert "nb_a_finaliser" in n
    # cam allees expose fix_plan/fix_reel/fix_delta if any
    cam_allees = (j.get("cam") or {}).get("allees") or []
    if cam_allees:
        for c in cam_allees:
            for k in ("fix_plan", "fix_reel", "fix_delta"):
                assert k in c


# ---------- Justification obligatoire écart >5% ----------
def test_justif_required_on_validate_when_gap_gt_5pct():
    # Set reel=80 (plan=100, écart 20%)
    r = requests.patch(f"{T}/allee",
                       json={"uid": UID, "products": [{"designation": "ES 1.5 noir", "reel": 80}]})
    assert r.status_code == 200, r.text
    # Clear any existing justification first via a status=a_faire + justification=""
    requests.patch(f"{T}/allee", json={"uid": UID, "status": "a_faire", "justification": ""})
    # Validate WITHOUT justification -> 400
    r = requests.patch(f"{T}/allee", json={"uid": UID, "status": "validee"})
    assert r.status_code == 400, f"expected 400, got {r.status_code}: {r.text}"
    assert "ustification" in r.text  # 'Justification requise'
    # Validate WITH justification -> 200
    r = requests.patch(f"{T}/allee",
                       json={"uid": UID, "status": "validee",
                             "justification": "Rayon réduit par le magasin"})
    assert r.status_code == 200
    j = requests.get(T).json()
    a = next(x for x in j["allees"] if x["uid"] == UID)
    assert a["status"] == "validee"
    assert a["justification"] == "Rayon réduit par le magasin"
    assert any(p["designation"] == "ES 1.5 noir" for p in a["justif_products"])


# ---------- a_finaliser status + night in red ----------
def test_a_finaliser_flags_night():
    r = requests.patch(f"{T}/allee", json={"uid": UID, "status": "a_finaliser"})
    assert r.status_code == 200
    j = requests.get(T).json()
    a = next(x for x in j["allees"] if x["uid"] == UID)
    assert a["status"] == "a_finaliser"
    n = next(n for n in j["nights"] if n["nuit"] == NIGHT)
    assert n["nb_a_finaliser"] >= 1
    # alert of type a_finaliser
    alerts = j.get("alerts") or []
    assert any(al.get("type") == "a_finaliser" for al in alerts), \
        f"no a_finaliser alert. types={[a.get('type') for a in alerts]}"


# ---------- extra_products persistence ----------
def test_extra_products_persist_and_show_in_stock():
    r = requests.patch(f"{T}/allee", json={
        "uid": UID,
        "extra_products": [{"designation": "ES 1.5 blanc", "qty": 12}],
    })
    assert r.status_code == 200
    j = requests.get(T).json()
    a = next(x for x in j["allees"] if x["uid"] == UID)
    extras = a.get("extra_products") or []
    assert any(e["designation"] == "ES 1.5 blanc" and e["qty"] == 12 for e in extras), extras
    # extras also aggregated somewhere in state (stock or extras section)
    payload_str = str(j)
    assert "ES 1.5 blanc" in payload_str


# ---------- fixations_reel on cam allée ----------
def test_cam_fixations_reel_persists():
    j = requests.get(T).json()
    cam_allees = (j.get("cam") or {}).get("allees") or []
    if not cam_allees:
        pytest.skip("no cam allée phased")
    cam_uid = cam_allees[0]["uid"]
    r = requests.patch(f"{T}/allee-cam", json={"uid": cam_uid, "fixations_reel": 3})
    assert r.status_code == 200, r.text
    j2 = requests.get(T).json()
    c = next(c for c in j2["cam"]["allees"] if c["uid"] == cam_uid)
    assert c["fix_reel"] == 3


# ---------- Excel rapport-nuit 3 sheets ----------
def test_rapport_nuit_xlsx_3_sheets():
    r = requests.get(f"{T}/rapport-nuit/{NIGHT}")
    assert r.status_code == 200, r.text
    assert "spreadsheet" in r.headers.get("content-type", "") or r.content[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=False)
    names = wb.sheetnames
    assert f"Nuit {NIGHT}" in names
    assert "Détail produits" in names
    assert "Synthèse déploiement" in names

    # Sheet Nuit N: contains KPI "Allées à finaliser une autre nuit" and Justification column
    ws1 = wb[f"Nuit {NIGHT}"]
    txt1 = "\n".join(str(c.value) for row in ws1.iter_rows() for c in row if c.value)
    assert "Allées à finaliser une autre nuit" in txt1
    assert "Justification" in txt1

    # Sheet Synthèse: KPI globaux + tableau nuits avec colonne "À finaliser"
    ws3 = wb["Synthèse déploiement"]
    txt3 = "\n".join(str(c.value) for row in ws3.iter_rows() for c in row if c.value)
    assert "Synthèse du déploiement" in txt3
    assert "À finaliser" in txt3
    assert "Nuits terminées" in txt3


# ---------- Admin regression ----------
def test_admin_routes_still_work(auth):
    r = auth.get(f"{BASE_URL}/api/suivi/{UPLOAD_ID}")
    assert r.status_code == 200
    assert r.json().get("allees")
    r = auth.patch(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/allee",
                   json={"uid": UID, "comment": "regression admin ok"})
    assert r.status_code == 200
    r = auth.get(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/rapport-nuit/{NIGHT}")
    assert r.status_code == 200
    assert r.content[:2] == b"PK"
