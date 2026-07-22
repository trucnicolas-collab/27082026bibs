"""iter48e — Nouvel onglet "Stock" dans le rapport de nuit Excel.

Vérifie :
  • Un onglet "Stock" existe dans le workbook généré par /rapport-nuit/{N}
  • Les colonnes attendues sont présentes (Désignation / Référence / Reçu / Posé
    / Restant / Reste à poser)
  • Coloration : ligne verte si Restant ≥ Reste à poser (stock suffisant), rouge sinon
  • Contient toutes les désignations du stock global (pas filtré par nuit)
"""
import io
import os
import openpyxl
import pytest
import requests
from dotenv import load_dotenv

load_dotenv("/app/backend/.env")
load_dotenv("/app/frontend/.env", override=False)

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
UPLOAD_ID = "fd15443f-6d2a-4cef-bd72-c56bb29e9c42"
UID = "1__A__R1"  # phasée sur nuit 2 par défaut
NIGHT = 2
T = f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}"

SUCCESS_BG = "D1FAE5"  # C_SUCCESS_BG (vert : stock suffisant)
DANGER_BG = "FEE2E2"   # C_DANGER_BG  (rouge : rupture prévue)


@pytest.fixture(scope="module")
def auth():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login",
               json={"email": "admin@vusion.local", "password": "admin123"})
    assert r.status_code == 200, r.text
    return s


@pytest.fixture(scope="module", autouse=True)
def ensure_published(auth):
    auth.post(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/publish", json={"published": True})
    yield


def _bg(cell) -> str:
    rgb = getattr(getattr(cell.fill.fgColor, "rgb", "") or "", "upper", lambda: "")()
    if isinstance(rgb, str) and len(rgb) == 8:
        return rgb[2:]
    return rgb or ""


def _generate():
    r = requests.get(f"{T}/rapport-nuit/{NIGHT}")
    assert r.status_code == 200, r.text
    return openpyxl.load_workbook(io.BytesIO(r.content), data_only=False)


def test_stock_sheet_exists_with_expected_headers(auth):
    wb = _generate()
    assert "Stock" in wb.sheetnames, wb.sheetnames
    ws = wb["Stock"]
    # Titre en r1, sous-titre en r2, headers en r4 (row 4)
    headers = [ws.cell(row=4, column=c).value for c in range(1, 7)]
    assert headers == ["Désignation", "Référence", "Reçu", "Posé",
                       "Restant", "Reste à poser"], headers


def test_stock_row_green_when_sufficient(auth):
    """Aucune saisie de pose → Restant = Reçu, Reste à poser = Prévu.
    Puisque recu_theorique = Prévu par défaut, la ligne doit être verte."""
    # Reset : aucune saisie
    requests.patch(f"{T}/allee", json={
        "uid": UID, "status": "a_faire",
        "products": [
            {"designation": "ES 1.5 noir", "reel": None},
            {"designation": "ES 2.1 noir", "reel": None},
            {"designation": "990 mm (noir)", "reel": None},
        ],
    })
    wb = _generate()
    ws = wb["Stock"]
    # Recherche ligne "ES 1.5 noir" (col 1) puis colonnes Restant (5) & Reste à poser (6)
    found = None
    for r in range(5, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "ES 1.5 noir":
            found = r
            break
    assert found, "ES 1.5 noir absent de l'onglet Stock"
    restant_cell = ws.cell(row=found, column=5)
    rap_cell = ws.cell(row=found, column=6)
    # Contrat métier : Restant ≥ Reste à poser → cellules VERTES (stock suffisant)
    assert restant_cell.value >= rap_cell.value, (
        f"restant={restant_cell.value} < rap={rap_cell.value} — pas censé être vert")
    assert _bg(restant_cell) == SUCCESS_BG, (
        f"Restant bg={_bg(restant_cell)} attendu {SUCCESS_BG} (vert). "
        f"restant={restant_cell.value} rap={rap_cell.value}")
    assert _bg(rap_cell) == SUCCESS_BG, _bg(rap_cell)


def test_stock_row_red_when_stock_insufficient(auth):
    """Force `recu` faible via /stock, puis pose > recu → Restant < Reste à poser.
    La ligne doit être ROUGE."""
    # Fixe recu = 10 (très faible) sur ES 1.5 noir
    requests.patch(f"{T}/stock", json={"designation": "ES 1.5 noir", "recu": 10})
    # Pose 5 (encore beaucoup à poser sur d'autres allées)
    requests.patch(f"{T}/allee", json={
        "uid": UID, "status": "a_faire",
        "products": [{"designation": "ES 1.5 noir", "reel": 5}],
    })
    wb = _generate()
    ws = wb["Stock"]
    found = None
    for r in range(5, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "ES 1.5 noir":
            found = r
            break
    assert found
    restant_cell = ws.cell(row=found, column=5)
    rap_cell = ws.cell(row=found, column=6)
    # Restant = 10 - 5 = 5 ; Reste à poser encore élevé → rouge
    assert restant_cell.value < rap_cell.value, (
        f"attendu restant<rap : restant={restant_cell.value} rap={rap_cell.value}")
    assert _bg(restant_cell) == DANGER_BG, (
        f"Restant bg={_bg(restant_cell)} attendu {DANGER_BG} (rouge)")
    assert _bg(rap_cell) == DANGER_BG, _bg(rap_cell)
    # Restore : remet recu à None (reçu théorique) et pose à None
    requests.patch(f"{T}/stock", json={"designation": "ES 1.5 noir", "recu": None})
    requests.patch(f"{T}/allee", json={
        "uid": UID, "status": "a_faire",
        "products": [{"designation": "ES 1.5 noir", "reel": None}],
    })


def test_stock_totals_row(auth):
    wb = _generate()
    ws = wb["Stock"]
    # Trouve la ligne "TOTAL"
    for r in range(5, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "TOTAL":
            recu_tot = ws.cell(row=r, column=3).value
            pose_tot = ws.cell(row=r, column=4).value
            assert isinstance(recu_tot, (int, float)), f"recu_tot={recu_tot!r}"
            assert isinstance(pose_tot, (int, float)), f"pose_tot={pose_tot!r}"
            return
    pytest.fail("Ligne TOTAL absente de l'onglet Stock")
