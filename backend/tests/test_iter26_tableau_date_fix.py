"""Iter26 — Fix Tableau date (Excel + PPTX slide 11).

Bug 1 : ligne « EEG ES+SA » affichait seulement l'ES pur (label trompeur).
Bug 2 : ligne « SA magasin » affichait le total SA du node (incl. SA à installer)
        au lieu du reliquat magasin (SA total - SA à installer).

Fix : `totals_by_nuit[n]["eeg"] = ES pur + bonus + flèches + SA à installer`
      `totals_by_nuit[n]["sa"]  = max(0, SA total node - SA à installer)`
"""
import io
import os
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
API = f"{BASE_URL}/api"

ADMIN_EMAIL = "admin@vusion.local"
ADMIN_PWD = "admin123"
DATASET_ID = "fd15443f-6d2a-4cef-bd72-c56bb29e9c42"


def _session():
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PWD})
    assert r.status_code == 200, r.text
    return s


def test_tableau_date_eeg_es_sa_includes_sa_a_installer():
    """Ligne « EEG ES+SA » du Tableau date = ES pur + bonus + flèches + SA à installer.

    Dataset test : 1 allée Nuit 2 avec es_15=105, es_21=50 (bonus=0, fleches=0)
    et SA à installer=60. Attendu : 155 + 60 = 215.
    """
    s = _session()
    r = s.get(f"{API}/export/{DATASET_ID}")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Tableau date"]
    rows = list(ws.iter_rows(values_only=True))
    header_row = next(r for r in rows if r and r[1] == "Nuit 1")
    idx_n2 = header_row.index("Nuit 2")
    eeg_row = next(r for r in rows if r and r[0] == "EEG ES+SA")
    assert eeg_row[idx_n2] == 215, \
        f"EEG ES+SA Nuit 2 attendu 215 (155 ES + 60 SA à installer), obtenu {eeg_row[idx_n2]}"


def test_tableau_date_sa_magasin_excludes_sa_a_installer():
    """Ligne « SA magasin » = SA total du node moins SA à installer.

    Dataset test : SA à installer = 60, SA total du node = 60 → SA magasin = 0.
    """
    s = _session()
    r = s.get(f"{API}/export/{DATASET_ID}")
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Tableau date"]
    rows = list(ws.iter_rows(values_only=True))
    header_row = next(r for r in rows if r and r[1] == "Nuit 1")
    idx_n2 = header_row.index("Nuit 2")
    sa_row = next(r for r in rows if r and r[0] == "SA magasin")
    assert (sa_row[idx_n2] or 0) == 0, \
        f"SA magasin Nuit 2 attendu 0 (aucun SA hors phasage), obtenu {sa_row[idx_n2]}"


def test_tableau_date_shows_all_nights_including_last():
    """Le Tableau date doit lister TOUTES les nuits ES + Cam sans troncature."""
    s = _session()
    r = s.get(f"{API}/export/{DATASET_ID}")
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Tableau date"]
    rows = list(ws.iter_rows(values_only=True))
    # Récupère tous les headers Nuit X présents
    all_night_headers = []
    for row in rows:
        if not row:
            continue
        for c in row:
            if isinstance(c, str) and c.startswith("Nuit "):
                all_night_headers.append(c)
    # Au minimum on doit voir chaque nuit du dataset une fois
    unique = set(all_night_headers)
    # Le dataset test a nb_es=10 nuits ES + 1 nuit Cam
    assert "Nuit 1" in unique
    assert "Nuit 2" in unique
    # Vérifie la dernière nuit (Nuit 10 minimum)
    numbers = sorted({int(h.split()[1]) for h in unique})
    assert max(numbers) >= 10, f"Dernière nuit visible={max(numbers)}, attendu >= 10"
