"""
Backend tests for Excel Inventory API.
Covers: upload-excel, dataset, export, error handling.
"""
import io
import os
import pytest
import requests
import openpyxl

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')
if not BASE_URL:
    # fallback to frontend .env
    try:
        with open('/app/frontend/.env') as f:
            for line in f:
                if line.startswith('REACT_APP_BACKEND_URL='):
                    BASE_URL = line.split('=', 1)[1].strip().rstrip('/')
    except Exception:
        pass

SAMPLE_PATH = "/tmp/sample.xlsx"


@pytest.fixture(scope="module")
def upload_response():
    """Upload sample file once, share across tests."""
    assert os.path.exists(SAMPLE_PATH), f"Sample file missing at {SAMPLE_PATH}"
    with open(SAMPLE_PATH, "rb") as f:
        files = {"file": ("sample.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/upload-excel", files=files, timeout=180)
    assert r.status_code == 200, f"Upload failed: {r.status_code} - {r.text[:500]}"
    return r.json()


# ---------- Health ----------
class TestHealth:
    def test_root(self):
        r = requests.get(f"{BASE_URL}/api/", timeout=30)
        assert r.status_code == 200
        data = r.json()
        assert "message" in data


# ---------- Upload Excel ----------
class TestUploadExcel:
    def test_upload_returns_required_fields(self, upload_response):
        d = upload_response
        assert "upload_id" in d and isinstance(d["upload_id"], str) and len(d["upload_id"]) > 0
        assert "row_count" in d
        assert "data" in d
        assert "recap" in d["data"]
        assert "secteur" in d["data"]
        assert "raw" in d["data"]

    def test_row_count_19780(self, upload_response):
        assert upload_response["row_count"] == 19780, f"Expected 19780 rows, got {upload_response['row_count']}"

    def test_raw_count_matches(self, upload_response):
        assert len(upload_response["data"]["raw"]) == 19780

    def test_recap_total_eeg_76366(self, upload_response):
        recap = upload_response["data"]["recap"]
        eeg_total = [r for r in recap if r["kind"] == "header" and r["type"] == "EEG"]
        assert len(eeg_total) == 1, f"Expected 1 TOTAL EEG header, got {len(eeg_total)}"
        assert eeg_total[0]["quantite"] == 76366, f"Expected EEG total 76366, got {eeg_total[0]['quantite']}"

    def test_recap_spare_lines(self, upload_response):
        recap = upload_response["data"]["recap"]
        # Une ligne Spare est désormais générée après chaque produit (kind='product')
        products = [r for r in recap if r["kind"] == "product"]
        spares = [r for r in recap if r["kind"] == "spare"]
        assert len(spares) == len(products), f"Expected one spare per product: {len(products)} products vs {len(spares)} spares"

        # Vérification : la somme des spares EEG = somme ceil(qty*0.05) par produit EEG
        import math
        eeg_products = [r for r in recap if r["kind"] == "product" and r["type"] == "EEG"]
        expected_eeg_total = sum(math.ceil(p["quantite"] * 0.05) for p in eeg_products)
        eeg_spares = [r for r in recap if r["kind"] == "spare" and r["type"] == "EEG"]
        actual_eeg_total = sum(s["quantite"] for s in eeg_spares)
        assert actual_eeg_total == expected_eeg_total, f"EEG spares sum {actual_eeg_total} != expected {expected_eeg_total}"

        # Pour chaque spare on doit avoir une référence non vide (héritée du produit)
        for s in spares:
            assert s["designation"].startswith("Spare (+5%)"), f"Bad designation: {s['designation']}"

    def test_recap_inclineur_rail_9669(self, upload_response):
        recap = upload_response["data"]["recap"]
        inclineurs = [r for r in recap if r["kind"] == "inclineur"]
        assert len(inclineurs) == 1, f"Expected 1 inclineur line, got {len(inclineurs)}"
        assert inclineurs[0]["type"].lower() == "rail"
        assert inclineurs[0]["quantite"] == 9669, f"Expected inclineur 9669, got {inclineurs[0]['quantite']}"

    def test_recap_three_empty_lines(self, upload_response):
        recap = upload_response["data"]["recap"]
        empties = [r for r in recap if r["kind"] == "empty"]
        assert len(empties) == 3, f"Expected 3 empty lines, got {len(empties)}"
        # Empty lines should be at end of recap
        assert all(r["kind"] == "empty" for r in recap[-3:]), "Empty lines should be at end"

    def test_secteur_rows_154(self, upload_response):
        secteur = upload_response["data"]["secteur"]
        assert len(secteur) == 154, f"Expected 154 secteur rows, got {len(secteur)}"

    def test_secteur_fields_present(self, upload_response):
        secteur = upload_response["data"]["secteur"]
        required = {"secteur", "rayon", "allee", "nb_eeg_es", "nb_eeg_sa", "nb_rail", "nb_camera"}
        for row in secteur[:5]:
            missing = required - set(row.keys())
            assert not missing, f"Missing fields in secteur row: {missing}"


# ---------- Get Dataset ----------
class TestDataset:
    def test_get_dataset(self, upload_response):
        upload_id = upload_response["upload_id"]
        r = requests.get(f"{BASE_URL}/api/dataset/{upload_id}", timeout=60)
        assert r.status_code == 200
        d = r.json()
        assert d["upload_id"] == upload_id
        assert "data" in d
        assert len(d["data"]["raw"]) == 19780

    def test_get_dataset_not_found(self):
        r = requests.get(f"{BASE_URL}/api/dataset/nonexistent-id-xxx", timeout=30)
        assert r.status_code == 404


# ---------- Export ----------
class TestExport:
    def test_export_all_three_sheets(self, upload_response):
        upload_id = upload_response["upload_id"]
        r = requests.get(f"{BASE_URL}/api/export/{upload_id}?sheet=all", timeout=180)
        assert r.status_code == 200
        ct = r.headers.get("content-type", "")
        assert "spreadsheet" in ct or "octet-stream" in ct, f"Bad content-type: {ct}"
        # Verify it's a valid xlsx with 3 sheets
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
        sheet_names = wb.sheetnames
        assert "Données" in sheet_names, f"Missing 'Données' sheet: {sheet_names}"
        assert "Récapitulatif" in sheet_names, f"Missing 'Récapitulatif' sheet: {sheet_names}"
        assert "Par Secteur" in sheet_names, f"Missing 'Par Secteur' sheet: {sheet_names}"
        wb.close()

    def test_export_not_found(self):
        r = requests.get(f"{BASE_URL}/api/export/nonexistent-id-xxx", timeout=30)
        assert r.status_code == 404


# ---------- Error handling ----------
class TestErrorHandling:
    def test_upload_non_excel_returns_400(self):
        files = {"file": ("test.txt", b"plain text content", "text/plain")}
        r = requests.post(f"{BASE_URL}/api/upload-excel", files=files, timeout=30)
        assert r.status_code == 400, f"Expected 400, got {r.status_code}: {r.text[:200]}"

    def test_upload_excel_missing_columns_returns_400(self):
        # Build a valid xlsx that lacks required columns
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["foo", "bar", "baz"])
        ws.append([1, 2, 3])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        files = {"file": ("bad.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/upload-excel", files=files, timeout=60)
        assert r.status_code == 400, f"Expected 400, got {r.status_code}: {r.text[:200]}"
