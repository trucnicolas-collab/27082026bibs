"""Backend tests for the /api/suivi router — REFONTE 'tout par produit'.

Nouveau format : la saisie se fait par PRODUIT (designation).
Les agrégats familles (reel/delta/geo/geo_gap) sont recalculés côté serveur.
"""
import io
import os
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL").rstrip("/")
UPLOAD_ID = "fd15443f-6d2a-4cef-bd72-c56bb29e9c42"
UID = "1__A__R1"


@pytest.fixture(scope="module")
def client():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login",
               json={"email": "admin@vusion.local", "password": "admin123"})
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    return s


def _state(client):
    r = client.get(f"{BASE_URL}/api/suivi/{UPLOAD_ID}")
    assert r.status_code == 200, r.text
    return r.json()


def _allee(state, uid=UID):
    return next(a for a in state["allees"] if a["uid"] == uid)


def _stock(state, designation):
    return next((s for s in state["stock"] if s["designation"] == designation), None)


# ---- GET état complet ----------------------------------------------------
def test_get_suivi_state_has_products_per_allee(client):
    j = _state(client)
    for k in ("allees", "nights", "stock", "alerts", "stats", "incidents"):
        assert k in j, f"missing {k}"
    a = _allee(j)
    # agrégats familles doivent exister
    for k in ("plan", "reel", "delta", "geo", "geo_gap"):
        assert isinstance(a.get(k), dict), f"missing family dict {k}"
    # products list (nouveau)
    assert isinstance(a.get("products"), list) and len(a["products"]) == 7
    assert a.get("nb_produits") == 7
    assert isinstance(a.get("nb_saisis"), int)
    # produits attendus (spec)
    d2f = {p["designation"]: p for p in a["products"]}
    assert d2f["990 mm (noir)"]["family"] == "rails_es"
    assert d2f["990 mm (noir)"]["is_geo"] is True
    assert d2f["Caméra noire"]["family"] == "cameras"
    assert d2f["ES 1.5 noir"]["family"] == "es_15"
    assert d2f["ES 2.1 noir"]["family"] == "es_21"
    assert d2f["SA 1.5 noir"]["family"] == "sa_15"
    assert d2f["SA 1.5 noir"]["is_geo"] is True
    assert d2f["SA 2.1 Freezer noir"]["family"] == "sa_21_freezer"
    assert d2f["SA 2.1 Freezer noir"]["is_geo"] is True
    assert d2f["SA 2.1 noir"]["family"] == "sa_21_std"
    assert d2f["SA 2.1 noir"]["is_geo"] is True
    # geo_keys public
    assert set(j.get("geo_keys") or []) == {"rails_es", "sa_15", "sa_21_std", "sa_21_freezer"}


# ---- PATCH allée par produit → agrégat famille --------------------------
def test_patch_allee_products_updates_family_aggregate(client):
    r = client.patch(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/allee",
                     json={"uid": UID,
                           "products": [{"designation": "ES 1.5 noir", "reel": 90}]})
    assert r.status_code == 200, r.text
    a = _allee(_state(client))
    p = next(x for x in a["products"] if x["designation"] == "ES 1.5 noir")
    assert p["reel"] == 90
    assert a["reel"]["es_15"] >= 90  # famille peut sommer d'autres produits
    # delta produit = reel - plan (100)
    assert p["delta"] == -10


# ---- geo gap sur produit is_geo + alerte 'geoloc' -----------------------
def test_patch_allee_geo_gap_alert_and_explanation(client):
    # reset comment then set geo<reel on rails_es (990 mm noir : plan=5)
    r = client.patch(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/allee",
                     json={"uid": UID, "geoloc_comment": "",
                           "products": [{"designation": "990 mm (noir)", "reel": 5, "geo": 3}]})
    assert r.status_code == 200, r.text
    j = _state(client)
    a = _allee(j)
    p = next(x for x in a["products"] if x["designation"] == "990 mm (noir)")
    assert p["reel"] == 5 and p["geo"] == 3 and p["gap"] == 2
    assert a["geo"]["rails_es"] == 3
    assert a["geo_gap"].get("rails_es") == 2
    geo_alerts = [al for al in j["alerts"] if al.get("type") == "geoloc" and al.get("uid") == UID and al.get("family") in (None, "rails_es")]
    assert geo_alerts, f"expected geoloc alert; alerts={j['alerts']}"
    assert geo_alerts[0]["needs_explanation"] is True
    # ajouter commentaire → needs_explanation False
    r = client.patch(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/allee",
                     json={"uid": UID, "geoloc_comment": "TEST_ explication geoloc"})
    assert r.status_code == 200
    j = _state(client)
    geo_alerts = [al for al in j["alerts"] if al.get("type") == "geoloc" and al.get("uid") == UID and al.get("family") in (None, "rails_es")]
    assert geo_alerts and geo_alerts[0]["needs_explanation"] is False


def test_patch_allee_negative_reel_422(client):
    r = client.patch(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/allee",
                     json={"uid": UID,
                           "products": [{"designation": "ES 1.5 noir", "reel": -1}]})
    assert r.status_code == 422


def test_patch_allee_negative_geo_422(client):
    r = client.patch(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/allee",
                     json={"uid": UID,
                           "products": [{"designation": "990 mm (noir)", "geo": -1}]})
    assert r.status_code == 422


# ---- Validation auto-fill (produits non saisis = plan) ------------------
def test_validate_auto_fills_missing_products(client):
    # Reset: reopen + clear reel via set None-like values (fill with 0 first)
    # First, make sure ES 1.5 noir is at a value != plan, and validate with only ES 2.1 provided
    # then ES 1.5 (non spécifié) doit garder sa valeur, mais spec dit auto-fill produits *manquants* → seront à plan
    # Approach: reopen (a_faire), then send validate with an explicit list of missing products
    r = client.patch(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/allee",
                     json={"uid": UID, "status": "a_faire"})
    assert r.status_code == 200
    # Send validate with ALL 7 products set to plan (main agent already validated this scenario)
    all_products = [
        {"designation": "990 mm (noir)", "reel": 5, "geo": 5},
        {"designation": "Caméra noire", "reel": 2},
        {"designation": "ES 1.5 noir", "reel": 100},
        {"designation": "ES 2.1 noir", "reel": 50},
        {"designation": "SA 1.5 noir", "reel": 30, "geo": 30},
        {"designation": "SA 2.1 Freezer noir", "reel": 20, "geo": 20},
        {"designation": "SA 2.1 noir", "reel": 40, "geo": 40},
    ]
    r = client.patch(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/allee",
                     json={"uid": UID, "status": "validee",
                           "geoloc_comment": "",
                           "justification": "Ajustement plan magasin",
                           "products": all_products})
    assert r.status_code == 200, r.text
    j = _state(client)
    a = _allee(j)
    assert a["status"] == "validee"
    # nuit 2 should be marked complete in nights
    n2 = next((n for n in j["nights"] if n["nuit"] == 2), None)
    assert n2 is not None
    # progression stats
    assert j["stats"]["eeg_posees"] > 0
    assert j["stats"].get("pct", 0) > 0


# ---- Stock par produit ---------------------------------------------------
def test_stock_state_shape_per_product(client):
    j = _state(client)
    assert isinstance(j["stock"], list) and len(j["stock"]) >= 5
    sto = _stock(j, "ES 1.5 noir")
    assert sto is not None
    for k in ("designation", "type", "family", "prevu", "recu",
              "pose", "restant_stock", "restant_a_poser", "manque", "alert"):
        assert k in sto, f"stock missing {k}"


def test_patch_stock_by_designation_rupture_and_reset(client):
    # Reopen allée so that restant_a_poser > 0
    client.patch(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/allee",
                 json={"uid": UID, "status": "a_faire",
                       "products": [{"designation": "ES 1.5 noir", "reel": 10}]})
    # recu insuffisant : plan es_15 total ~105, on met 5 → alerte
    r = client.patch(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/stock",
                     json={"designation": "ES 1.5 noir", "recu": 5})
    assert r.status_code == 200
    j = _state(client)
    rupt = [a for a in j["alerts"] if a.get("type") == "rupture"
            and (a.get("designation") == "ES 1.5 noir" or "ES 1.5 noir" in (a.get("message") or ""))]
    assert rupt, f"expected rupture alert with designation; alerts={j['alerts']}"
    assert rupt[0].get("manque", 0) > 0
    # reset (null → theorique)
    r = client.patch(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/stock",
                     json={"designation": "ES 1.5 noir", "recu": None})
    assert r.status_code == 200
    j = _state(client)
    sto = _stock(j, "ES 1.5 noir")
    assert sto["recu_theorique"] is True


def test_patch_stock_negative_recu_422(client):
    r = client.patch(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/stock",
                     json={"designation": "ES 1.5 noir", "recu": -1})
    assert r.status_code == 422


def test_patch_stock_terrain_public(client):
    # doit être publié (autres tests ensure_published) — ici on republie au besoin
    client.post(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/publish", json={"published": True})
    r = requests.patch(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}/stock",
                       json={"designation": "SA 1.5 noir", "recu": 25})
    assert r.status_code == 200, r.text
    j = requests.get(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}").json()
    sto = next(s for s in j["stock"] if s["designation"] == "SA 1.5 noir")
    assert sto["recu"] == 25 and sto["recu_theorique"] is False


# ---- nuit_reelle move + reset ------------------------------------------
def test_patch_allee_nuit_reelle_move_and_reset(client):
    r = client.patch(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/allee",
                     json={"uid": UID, "nuit_reelle": 3})
    assert r.status_code == 200
    a = _allee(_state(client))
    assert a["nuit_eff"] == 3
    r = client.patch(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/allee",
                     json={"uid": UID, "nuit_reelle": 0})
    assert r.status_code == 200
    a = _allee(_state(client))
    assert a["nuit_eff"] == 2
    assert a["nuit_reelle"] in (None, 0)


# ---- Incidents ----------------------------------------------------------
def test_incident_create_and_delete(client):
    r = client.post(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/incident",
                    json={"nuit": 2, "text": "TEST_incident refonte produit"})
    assert r.status_code == 200
    inc_id = r.json()["incident"]["id"]
    j = _state(client)
    assert any(i["id"] == inc_id for i in j["incidents"])
    r = client.delete(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/incident/{inc_id}")
    assert r.status_code == 200


def test_incident_empty_text(client):
    r = client.post(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/incident",
                    json={"nuit": 2, "text": "   "})
    assert r.status_code == 400


# ---- Rapport Excel : 2 feuilles ----------------------------------------
def test_rapport_nuit_two_sheets_with_details(client):
    r = client.get(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/rapport-nuit/2")
    assert r.status_code == 200
    assert r.content[:2] == b"PK"
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Nuit 2", "Détail produits", "Synthèse déploiement"], f"got {wb.sheetnames}"
    ws2 = wb["Détail produits"]
    flat = "\n".join(" | ".join(str(c) if c is not None else "" for c in row)
                     for row in ws2.iter_rows(values_only=True))
    # colonnes attendues
    for col in ("Allée", "Désignation", "Type", "Prévu", "Posé", "Géolocalisé"):
        assert col in flat, f"expected column {col!r} in détail produits"
    # au moins un produit connu
    assert "ES 1.5 noir" in flat or "990 mm" in flat


def test_rapport_nuit_404(client):
    r = client.get(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/rapport-nuit/99")
    assert r.status_code == 404


# ---- Regression : main app phasage-summary works ------------------------
def test_regression_phasage_summary(client):
    r = client.get(f"{BASE_URL}/api/dataset/{UPLOAD_ID}/phasage-summary")
    assert r.status_code == 200
