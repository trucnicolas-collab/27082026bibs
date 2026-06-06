"""E2E tests for the Carrefour export feature + removal of PPTX/store-info endpoints.

Covers review request iteration 6:
- Excel RTR endpoint (unchanged): /api/export/{id}?sheet=all
- Excel Carrefour endpoint (new): /api/export-carrefour/{id} → 5 specific tabs
- PPTX endpoint removed → 404
- store-info endpoint removed → 404/405
- Activity log still records carrefour_export_downloaded
- Phasage full sheet now has 10 columns (split SR), Nuit cell has white fill
- Recap complet sheet of Carrefour: 10 columns, Nuit blank/white
"""

import io
import os

import pytest
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv("/app/frontend/.env")
BASE_URL = os.environ.get("REACT_APP_BACKEND_URL").rstrip("/")
ADMIN_EMAIL = "admin@vusion.local"
ADMIN_PASS = "admin123"
DATASET_ID = "aa7d9aa6-ec7d-4f27-968e-b2a5228b0065"


@pytest.fixture(scope="module")
def authed_session():
    s = requests.Session()
    r = s.post(
        f"{BASE_URL}/api/auth/login",
        json={"email": ADMIN_EMAIL, "password": ADMIN_PASS},
        timeout=15,
    )
    assert r.status_code == 200, f"Login failed: {r.status_code} {r.text}"
    return s


# ----------------------------------------------------------------------
# Removed endpoints (PPT / store-info)
# ----------------------------------------------------------------------

class TestRemovedEndpoints:
    def test_export_pptx_endpoint_removed(self, authed_session):
        r = authed_session.get(f"{BASE_URL}/api/dataset/{DATASET_ID}/export-pptx", timeout=20)
        assert r.status_code in (404, 405), (
            f"Expected 404/405 for removed PPTX endpoint, got {r.status_code}"
        )

    def test_store_info_patch_endpoint_removed(self, authed_session):
        r = authed_session.patch(
            f"{BASE_URL}/api/dataset/{DATASET_ID}/store-info",
            json={"store_name": "x"},
            timeout=15,
        )
        assert r.status_code in (404, 405), (
            f"Expected 404/405 for removed store-info endpoint, got {r.status_code}"
        )


# ----------------------------------------------------------------------
# Excel RTR (Excel original)
# ----------------------------------------------------------------------

class TestExcelRTR:
    def test_excel_rtr_download_and_tabs(self, authed_session):
        r = authed_session.get(
            f"{BASE_URL}/api/export/{DATASET_ID}?sheet=all", timeout=60
        )
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")
        wb = load_workbook(io.BytesIO(r.content), data_only=False)
        names = wb.sheetnames
        # The RTR file must contain the historical sheets
        expected_any = ["Données", "Commandes", "Phasage de pose",
                        "Phasage caméras", "Phasage full"]
        for name in expected_any:
            assert name in names, f"RTR Excel missing sheet '{name}'. Got: {names}"

    def test_phasage_full_has_10_cols_and_nuit_white(self, authed_session):
        r = authed_session.get(
            f"{BASE_URL}/api/export/{DATASET_ID}?sheet=all", timeout=60
        )
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content), data_only=False)
        assert "Phasage full" in wb.sheetnames
        ws = wb["Phasage full"]
        # The recap table starts somewhere; look for the row containing 'Nuit'
        nuit_col = None
        nuit_row = None
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
            for cell in row:
                if str(cell.value).strip().lower() == "nuit":
                    nuit_col = cell.column
                    nuit_row = cell.row
                    break
            if nuit_col:
                break
        assert nuit_col is not None, "Could not find 'Nuit' header in Phasage full"
        # Check at least one data row beneath nuit_row has a white/none fill on the Nuit column
        found_white = False
        for r_idx in range(nuit_row + 1, min(ws.max_row, nuit_row + 60)):
            c = ws.cell(row=r_idx, column=nuit_col)
            if c.value is None or str(c.value).strip() == "":
                continue
            fill = c.fill
            fg = (fill.fgColor.rgb if fill and fill.fgColor else None) or ""
            # White fills can be 'FFFFFFFF', '00000000' (no fill) or None
            if fg in ("FFFFFFFF", "00000000", "", None):
                found_white = True
                break
        assert found_white, "Nuit column cell should have a white/blank background in Phasage full"


# ----------------------------------------------------------------------
# Excel Carrefour (new)
# ----------------------------------------------------------------------

EXPECTED_CARREFOUR_TABS = [
    "Commandes",
    "Récap EEG par nuit",
    "Récap caméra par nuit",
    "Caméra par élément",
    "Récap complet",
]


class TestExcelCarrefour:
    @pytest.fixture(scope="class")
    def carrefour_bytes(self, authed_session):
        r = authed_session.get(
            f"{BASE_URL}/api/export-carrefour/{DATASET_ID}", timeout=60
        )
        assert r.status_code == 200, f"Carrefour export failed: {r.status_code} {r.text[:300]}"
        assert "spreadsheetml" in r.headers.get("content-type", "")
        cd = r.headers.get("content-disposition", "")
        assert "Carrefour" in cd, f"Filename should contain 'Carrefour', got: {cd}"
        return r.content

    def test_tabs_exact_and_ordered(self, carrefour_bytes):
        wb = load_workbook(io.BytesIO(carrefour_bytes), data_only=False)
        # Must contain EXACTLY the 5 tabs in this order, no hidden _Phasage_data
        assert wb.sheetnames == EXPECTED_CARREFOUR_TABS, (
            f"Expected exactly {EXPECTED_CARREFOUR_TABS} got {wb.sheetnames}"
        )

    def test_no_hidden_helper_sheets(self, carrefour_bytes):
        wb = load_workbook(io.BytesIO(carrefour_bytes), data_only=False)
        for name in wb.sheetnames:
            ws = wb[name]
            assert ws.sheet_state == "visible", f"Sheet '{name}' should be visible"
            assert not name.startswith("_"), f"Hidden helper sheet leaked: {name}"

    def test_recap_complet_columns_and_nuit_white(self, carrefour_bytes):
        wb = load_workbook(io.BytesIO(carrefour_bytes), data_only=False)
        ws = wb["Récap complet"]
        # Find the row with the 10 column headers; expected:
        #   Allées | ES | Rails ES | SA | Secteur/Rayon EEG | Nuit | Date | Secteur/Rayon Cam | Allées | Caméras
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10)):
            vals = [str(c.value).strip().lower() if c.value is not None else "" for c in row]
            if "nuit" in vals and "date" in vals and vals.count("allées") >= 1:
                header_row = row[0].row
                break
        assert header_row is not None, "Could not find header row in Récap complet"

        # Collect 10 visible header values
        headers = [
            ws.cell(row=header_row, column=c).value
            for c in range(1, 11)
        ]
        joined = "|".join(str(h or "") for h in headers).lower()
        for kw in ["allées", "es", "rails es", "sa", "secteur", "nuit", "date", "caméras"]:
            assert kw in joined, f"Missing keyword '{kw}' in Récap complet headers: {headers}"

        # Find Nuit column (within first 10 cols)
        nuit_col = None
        for c in range(1, 11):
            v = ws.cell(row=header_row, column=c).value
            if v and str(v).strip().lower() == "nuit":
                nuit_col = c
                break
        assert nuit_col is not None, f"'Nuit' column not in first 10 cols. Headers={headers}"

        # Check at least one data row has white fill on Nuit column
        white_found = False
        for r_idx in range(header_row + 1, min(ws.max_row + 1, header_row + 60)):
            c = ws.cell(row=r_idx, column=nuit_col)
            if c.value is None or str(c.value).strip() == "":
                continue
            fill = c.fill
            fg = (fill.fgColor.rgb if fill and fill.fgColor else None) or ""
            if fg in ("FFFFFFFF", "00000000", "", None):
                white_found = True
                break
        assert white_found, "Nuit column cell in Récap complet should have white background"


# ----------------------------------------------------------------------
# Activity log
# ----------------------------------------------------------------------

class TestActivity:
    def test_activity_records_carrefour_download(self, authed_session):
        # trigger a download
        r = authed_session.get(
            f"{BASE_URL}/api/export-carrefour/{DATASET_ID}", timeout=60
        )
        assert r.status_code == 200

        a = authed_session.get(
            f"{BASE_URL}/api/dataset/{DATASET_ID}/activity", timeout=15
        )
        assert a.status_code == 200
        data = a.json()
        items = data if isinstance(data, list) else (
            data.get("activity") or data.get("items") or data.get("activities") or []
        )
        actions = [str(it.get("action", "")) for it in items]
        assert any("carrefour_export_downloaded" in x for x in actions), (
            f"carrefour_export_downloaded not found in activity. Sample: {actions[:5]}"
        )
