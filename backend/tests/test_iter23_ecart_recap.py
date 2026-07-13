"""Iter23 — Vue « Écart phasage vs réel » dans /materiel/{nuit}."""
import os
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
API = f"{BASE_URL}/api"

ADMIN_EMAIL = "admin@vusion.local"
ADMIN_PWD = "admin123"
DATASET_ID = "fd15443f-6d2a-4cef-bd72-c56bb29e9c42"


def _session():
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PWD})
    assert r.status_code == 200, r.text
    return s


def test_materiel_nuit_returns_ecarts_and_stats():
    s = _session()
    r = s.get(f"{API}/suivi/{DATASET_ID}/materiel/2", params={"mode": "eeg"})
    assert r.status_code == 200, r.text
    d = r.json()
    assert "ecarts" in d, "champ ecarts manquant"
    assert "ecart_stats" in d, "champ ecart_stats manquant"
    ec = d["ecart_stats"]
    for k in ("nb_saisis", "nb_conforme", "nb_bonus", "nb_manque", "complete"):
        assert k in ec, f"clé {k} manquante dans ecart_stats"
    assert ec["nb_saisis"] == ec["nb_conforme"] + ec["nb_bonus"] + ec["nb_manque"]


def test_ecart_status_classification():
    """Chaque écart doit avoir un status valide et cohérent avec delta/plan."""
    s = _session()
    r = s.get(f"{API}/suivi/{DATASET_ID}/materiel/2", params={"mode": "eeg"})
    d = r.json()
    for e in d.get("ecarts", []):
        assert e["status"] in ("conforme", "bonus", "manque")
        plan = float(e["plan"] or 0)
        delta = float(e["delta"] or 0)
        if plan > 0:
            pct = abs(delta) / plan
            if e["status"] == "conforme":
                assert pct <= 0.05 + 1e-9, f"conforme mais pct={pct} pour {e['designation']}"
            elif e["status"] == "bonus":
                assert delta > 0 and pct > 0.05
            elif e["status"] == "manque":
                assert delta < 0 and pct > 0.05


def test_allee_has_own_ecarts():
    """Chaque allée retournée doit avoir sa propre liste ecarts (drill-down)."""
    s = _session()
    r = s.get(f"{API}/suivi/{DATASET_ID}/materiel/2", params={"mode": "eeg"})
    d = r.json()
    for a in d["allees"]:
        assert "ecarts" in a
        assert "status" in a


def test_ecart_recap_cam_mode_no_crash():
    """Mode cam avec une allée cam Nuit 1 : la structure ecarts doit exister."""
    s = _session()
    r = s.get(f"{API}/suivi/{DATASET_ID}/materiel/1", params={"mode": "cam"})
    assert r.status_code == 200
    d = r.json()
    assert "ecarts" in d
    assert "ecart_stats" in d


def test_rapport_nuit_xlsx_contient_feuille_ecart():
    """Le rapport Excel d'une nuit avec des saisies contient la feuille 'Écart phasage vs réel'."""
    import io
    from openpyxl import load_workbook
    s = _session()
    # S'assurer qu'au moins un réel est saisi pour la nuit 2
    r0 = s.patch(f"{API}/suivi/{DATASET_ID}/allee",
                 json={"uid": "1__A__R1",
                       "products": [{"designation": "ES 1.5 noir", "reel": 95}]})
    assert r0.status_code == 200, r0.text
    r = s.get(f"{API}/suivi/{DATASET_ID}/rapport-nuit/2")
    assert r.status_code == 200, r.text
    assert r.headers.get("content-type", "").startswith("application/vnd.openxmlformats"), r.headers
    wb = load_workbook(io.BytesIO(r.content))
    assert "Écart phasage vs réel" in wb.sheetnames, f"Feuilles: {wb.sheetnames}"
    ws = wb["Écart phasage vs réel"]
    # Titre reconnaissable
    rows = list(ws.iter_rows(values_only=True, max_row=15))
    all_text = "\n".join(str(c) for r in rows for c in r if c is not None)
    assert "Écart phasage vs réel" in all_text
    assert "EEG · Écarts par produit" in all_text
    # Tableau des colonnes
    assert any("Désignation" == r[0] for r in rows if r[0])
    # Au moins une ligne TOTAL
    assert any(str(r[0]).strip() == "TOTAL" for r in rows if r and r[0])
