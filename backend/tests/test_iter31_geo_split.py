"""Iter31 — POSE vs GÉOLOCALISATION séparation.

Vérifie que :
- GEO_KEYS = [rails_es, sa_15, sa_21_std] (retiré sa_21_freezer)
- /api/suivi/{upload_id}/materiel/{nuit}?mode=eeg renvoie is_geo=true
  UNIQUEMENT pour rails_es / sa_15 / sa_21_std.
- Les items ecarts contiennent family, is_geo, geo (renseigné si is_geo).
- L'export Excel /rapport-nuit/{nuit} génère un fichier valide.
"""
import os
import io
import pytest
import requests
import openpyxl

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://go-lang-43.preview.emergentagent.com').rstrip('/')
DATASET_ID = "fd15443f-6d2a-4cef-bd72-c56bb29e9c42"
GEO_KEYS_EXPECTED = ["rails_es", "sa_15", "sa_21_std"]
NON_GEO_FAMILIES = ["es_15", "es_21", "sa_21_freezer", "sa_42", "fleches", "cameras"]


@pytest.fixture(scope="module")
def client():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    r = s.post(f"{BASE_URL}/api/auth/login",
               json={"email": "admin@vusion.local", "password": "admin123"}, timeout=15)
    assert r.status_code == 200, f"Login failed: {r.status_code} {r.text}"
    return s


def test_suivi_dataset_geo_keys(client):
    """Le champ state.geo_keys expose la classification GEO côté frontend."""
    r = client.get(f"{BASE_URL}/api/suivi/{DATASET_ID}", timeout=20)
    assert r.status_code == 200
    d = r.json()
    assert d["geo_keys"] == GEO_KEYS_EXPECTED, (
        f"GEO_KEYS incorrect: {d['geo_keys']} attendu {GEO_KEYS_EXPECTED}")


def test_materiel_nuit_sa15_is_geo(client):
    """Nuit 1 : SA 1.5 noir → is_geo=true, geo=7, family=sa_15."""
    r = client.get(f"{BASE_URL}/api/suivi/{DATASET_ID}/materiel/1?mode=eeg", timeout=20)
    assert r.status_code == 200
    body = r.json()
    ecarts = body.get("ecarts") or []
    sa15 = next((e for e in ecarts if e["designation"] == "SA 1.5 noir"), None)
    assert sa15 is not None, "SA 1.5 noir absent des écarts nuit 1"
    assert sa15["family"] == "sa_15"
    assert sa15["is_geo"] is True
    assert sa15["plan"] == 15
    assert sa15["reel"] == 12
    assert sa15["geo"] == 7


def test_materiel_nuit_es15_not_geo(client):
    """Nuit 1 : ES 1.5 noir → is_geo=false, geo=None, family=es_15."""
    r = client.get(f"{BASE_URL}/api/suivi/{DATASET_ID}/materiel/1?mode=eeg", timeout=20)
    body = r.json()
    ecarts = body.get("ecarts") or []
    es15 = next((e for e in ecarts if e["designation"] == "ES 1.5 noir"), None)
    assert es15 is not None, "ES 1.5 noir absent des écarts nuit 1"
    assert es15["family"] == "es_15"
    assert es15["is_geo"] is False
    assert es15["geo"] is None


def test_all_nights_family_classification(client):
    """Toutes les nuits : is_geo=true ⇔ family ∈ GEO_KEYS_EXPECTED."""
    for n in range(1, 11):
        r = client.get(f"{BASE_URL}/api/suivi/{DATASET_ID}/materiel/{n}?mode=eeg", timeout=15)
        if r.status_code == 404:
            continue
        assert r.status_code == 200
        for e in (r.json().get("ecarts") or []):
            fam = e.get("family")
            is_geo = e.get("is_geo")
            if fam in GEO_KEYS_EXPECTED:
                assert is_geo is True, f"Nuit {n} {e['designation']} fam={fam} devrait être is_geo=true"
            else:
                assert is_geo is False, f"Nuit {n} {e['designation']} fam={fam} ne doit PAS être is_geo=true"
                assert e.get("geo") is None, f"Nuit {n} {e['designation']} fam={fam} doit avoir geo=null"


def test_allee_level_ecarts_have_family_and_is_geo(client):
    """Les ecarts au niveau allée (drill-down) doivent aussi porter family/is_geo/geo."""
    r = client.get(f"{BASE_URL}/api/suivi/{DATASET_ID}/materiel/1?mode=eeg", timeout=20)
    assert r.status_code == 200
    body = r.json()
    for a in body.get("allees", []):
        for e in a.get("ecarts", []):
            assert "family" in e
            assert "is_geo" in e
            if e["is_geo"]:
                assert e["family"] in GEO_KEYS_EXPECTED


def test_rapport_nuit_excel_generation(client):
    """L'export Excel du rapport de nuit doit être un fichier XLSX valide et contenir
    les feuilles standards. NB : la nuit 1 de ce dataset n'a pas de caméras.
    On vérifie donc qu'aucune régression n'a été introduite."""
    r = client.get(f"{BASE_URL}/api/suivi/{DATASET_ID}/rapport-nuit/1", timeout=30)
    assert r.status_code == 200
    assert len(r.content) > 5000, "Excel report trop petit"
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "Nuit 1" in wb.sheetnames
    assert "Détail produits" in wb.sheetnames


def test_backend_regression_suivi_endpoint(client):
    """Régression : /api/suivi/{upload_id} répond bien avec la structure attendue."""
    r = client.get(f"{BASE_URL}/api/suivi/{DATASET_ID}", timeout=20)
    assert r.status_code == 200
    d = r.json()
    for key in ("upload_id", "nb_nuits", "allees", "nights", "stock", "alerts",
                "cam", "stats", "geo_keys"):
        assert key in d, f"Clé manquante dans /api/suivi: {key}"
