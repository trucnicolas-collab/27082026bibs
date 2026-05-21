"""
Backend tests for inline recap-row edition endpoints.
Covers: PATCH/POST/DELETE /api/dataset/{upload_id}/recap-row, quantite parsing, export reflection.
"""
import io
import os
import pytest
import requests
import openpyxl

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')
if not BASE_URL:
    try:
        with open('/app/frontend/.env') as f:
            for line in f:
                if line.startswith('REACT_APP_BACKEND_URL='):
                    BASE_URL = line.split('=', 1)[1].strip().rstrip('/')
    except Exception:
        pass

SAMPLE_PATH = "/tmp/sample.xlsx"


@pytest.fixture(scope="module")
def upload_id():
    """Upload sample once for this module of tests."""
    assert os.path.exists(SAMPLE_PATH), f"Sample file missing at {SAMPLE_PATH}"
    with open(SAMPLE_PATH, "rb") as f:
        files = {"file": ("sample.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/upload-excel", files=files, timeout=180)
    assert r.status_code == 200, f"Upload failed: {r.status_code} - {r.text[:500]}"
    return r.json()["upload_id"]


def _get_recap(uid):
    r = requests.get(f"{BASE_URL}/api/dataset/{uid}", timeout=60)
    assert r.status_code == 200, r.text[:300]
    return r.json()["data"]["recap"]


def _find_index_by_kind(recap, kind):
    for i, row in enumerate(recap):
        if row["kind"] == kind:
            return i
    return -1


# ---------- PATCH recap-row ----------
class TestPatchRecapRow:
    def test_patch_empty_to_manual(self, upload_id):
        recap = _get_recap(upload_id)
        idx = _find_index_by_kind(recap, "empty")
        assert idx >= 0, "No empty row available"

        payload = {"type": "Custom", "reference": "REF-001", "designation": "Manual added", "quantite": "42"}
        r = requests.patch(f"{BASE_URL}/api/dataset/{upload_id}/recap-row/{idx}", json=payload, timeout=30)
        assert r.status_code == 200, r.text[:300]
        data = r.json()
        assert data["index"] == idx
        row = data["row"]
        assert row["kind"] == "manual"
        assert row["type"] == "Custom"
        assert row["reference"] == "REF-001"
        assert row["designation"] == "Manual added"
        assert row["quantite"] == 42.0

        # Verify persistence
        recap2 = _get_recap(upload_id)
        assert recap2[idx]["kind"] == "manual"
        assert recap2[idx]["quantite"] == 42.0
        assert recap2[idx]["designation"] == "Manual added"

    def test_patch_manual_back_to_empty(self, upload_id):
        # First make a row manual, then clear it
        recap = _get_recap(upload_id)
        idx = _find_index_by_kind(recap, "empty")
        assert idx >= 0

        # Set to manual
        r = requests.patch(f"{BASE_URL}/api/dataset/{upload_id}/recap-row/{idx}",
                           json={"type": "T", "reference": "R", "designation": "D", "quantite": "5"}, timeout=30)
        assert r.status_code == 200
        assert r.json()["row"]["kind"] == "manual"

        # Clear it (all empty + qty 0)
        r = requests.patch(f"{BASE_URL}/api/dataset/{upload_id}/recap-row/{idx}",
                           json={"type": "", "reference": "", "designation": "", "quantite": "0"}, timeout=30)
        assert r.status_code == 200, r.text[:300]
        row = r.json()["row"]
        assert row["kind"] == "empty"

        # Also test with truly empty body fields
        r = requests.patch(f"{BASE_URL}/api/dataset/{upload_id}/recap-row/{idx}",
                           json={"type": "", "reference": "", "designation": "", "quantite": ""}, timeout=30)
        assert r.status_code == 200
        assert r.json()["row"]["kind"] == "empty"

    @pytest.mark.parametrize("kind_to_test", ["header", "inclineur", "product"])
    def test_patch_non_editable_returns_400(self, upload_id, kind_to_test):
        recap = _get_recap(upload_id)
        idx = _find_index_by_kind(recap, kind_to_test)
        assert idx >= 0, f"No row with kind={kind_to_test}"
        r = requests.patch(f"{BASE_URL}/api/dataset/{upload_id}/recap-row/{idx}",
                           json={"type": "X", "reference": "Y", "designation": "Z", "quantite": "1"}, timeout=30)
        assert r.status_code == 400, f"Expected 400 for {kind_to_test}, got {r.status_code}: {r.text[:200]}"

    def test_patch_unknown_upload_id_returns_404(self):
        r = requests.patch(f"{BASE_URL}/api/dataset/no-such-id/recap-row/0",
                           json={"type": "A"}, timeout=30)
        assert r.status_code == 404

    def test_patch_index_out_of_bounds_returns_404(self, upload_id):
        r = requests.patch(f"{BASE_URL}/api/dataset/{upload_id}/recap-row/999999",
                           json={"type": "A"}, timeout=30)
        assert r.status_code == 404
        r2 = requests.patch(f"{BASE_URL}/api/dataset/{upload_id}/recap-row/-1",
                            json={"type": "A"}, timeout=30)
        assert r2.status_code == 404


# ---------- POST add recap-row ----------
class TestAddRecapRow:
    def test_post_adds_empty_row_at_end(self, upload_id):
        recap_before = _get_recap(upload_id)
        n_before = len(recap_before)

        r = requests.post(f"{BASE_URL}/api/dataset/{upload_id}/recap-row", timeout=30)
        assert r.status_code == 200, r.text[:300]
        data = r.json()
        assert data["index"] == n_before  # appended at the end
        assert data["row"]["kind"] == "empty"
        assert data["row"]["type"] == ""
        assert data["row"]["reference"] == ""
        assert data["row"]["designation"] == ""
        assert data["row"]["quantite"] == ""

        # Verify persisted
        recap_after = _get_recap(upload_id)
        assert len(recap_after) == n_before + 1
        assert recap_after[-1]["kind"] == "empty"

    def test_post_unknown_upload_id_returns_404(self):
        r = requests.post(f"{BASE_URL}/api/dataset/no-such-id/recap-row", timeout=30)
        assert r.status_code == 404


# ---------- DELETE recap-row ----------
class TestDeleteRecapRow:
    def test_delete_empty_row(self, upload_id):
        # Add a fresh empty row to delete safely
        add = requests.post(f"{BASE_URL}/api/dataset/{upload_id}/recap-row", timeout=30)
        assert add.status_code == 200
        idx = add.json()["index"]

        recap_before = _get_recap(upload_id)
        n_before = len(recap_before)
        assert recap_before[idx]["kind"] == "empty"

        r = requests.delete(f"{BASE_URL}/api/dataset/{upload_id}/recap-row/{idx}", timeout=30)
        assert r.status_code == 200, r.text[:300]
        data = r.json()
        assert data["ok"] is True
        assert data["remaining"] == n_before - 1

        recap_after = _get_recap(upload_id)
        assert len(recap_after) == n_before - 1

    def test_delete_manual_row(self, upload_id):
        # Add then patch to manual, then delete
        add = requests.post(f"{BASE_URL}/api/dataset/{upload_id}/recap-row", timeout=30)
        idx = add.json()["index"]
        p = requests.patch(f"{BASE_URL}/api/dataset/{upload_id}/recap-row/{idx}",
                           json={"type": "M", "reference": "M1", "designation": "Manu", "quantite": "7"}, timeout=30)
        assert p.status_code == 200 and p.json()["row"]["kind"] == "manual"

        r = requests.delete(f"{BASE_URL}/api/dataset/{upload_id}/recap-row/{idx}", timeout=30)
        assert r.status_code == 200
        assert r.json()["ok"] is True

    @pytest.mark.parametrize("kind_to_test", ["header", "inclineur", "product"])
    def test_delete_non_deletable_returns_400(self, upload_id, kind_to_test):
        recap = _get_recap(upload_id)
        idx = _find_index_by_kind(recap, kind_to_test)
        assert idx >= 0
        r = requests.delete(f"{BASE_URL}/api/dataset/{upload_id}/recap-row/{idx}", timeout=30)
        assert r.status_code == 400, f"Expected 400 for {kind_to_test}, got {r.status_code}: {r.text[:200]}"

    def test_delete_invalid_index_returns_404(self, upload_id):
        r = requests.delete(f"{BASE_URL}/api/dataset/{upload_id}/recap-row/999999", timeout=30)
        assert r.status_code == 404
        r2 = requests.delete(f"{BASE_URL}/api/dataset/{upload_id}/recap-row/-5", timeout=30)
        assert r2.status_code == 404

    def test_delete_unknown_upload_id_returns_404(self):
        r = requests.delete(f"{BASE_URL}/api/dataset/no-such-id/recap-row/0", timeout=30)
        assert r.status_code == 404


# ---------- Quantite parsing ----------
class TestQuantiteParsing:
    @pytest.mark.parametrize("input_val,expected", [
        ("42", 42.0),
        ("42.5", 42.5),
        ("42,5", 42.5),
        ("1 000", 1000.0),
        (42, 42.0),
        (42.5, 42.5),
    ])
    def test_quantite_conversion(self, upload_id, input_val, expected):
        # use a fresh added row to avoid interfering with other tests
        add = requests.post(f"{BASE_URL}/api/dataset/{upload_id}/recap-row", timeout=30)
        idx = add.json()["index"]

        payload = {"type": "T", "reference": "R", "designation": "Q-test", "quantite": input_val}
        r = requests.patch(f"{BASE_URL}/api/dataset/{upload_id}/recap-row/{idx}", json=payload, timeout=30)
        assert r.status_code == 200, r.text[:300]
        row = r.json()["row"]
        assert row["quantite"] == expected, f"Input {input_val!r} -> got {row['quantite']!r}, expected {expected}"

        # cleanup
        requests.delete(f"{BASE_URL}/api/dataset/{upload_id}/recap-row/{idx}", timeout=30)


# ---------- Export reflects manual edits ----------
class TestExportReflectsEdits:
    def test_export_recap_contains_manual_row(self, upload_id):
        # Add and PATCH a manual row
        add = requests.post(f"{BASE_URL}/api/dataset/{upload_id}/recap-row", timeout=30)
        idx = add.json()["index"]
        marker_desig = "TEST_MANUAL_EXPORT_XYZ"
        payload = {"type": "MTYPE", "reference": "MREF99", "designation": marker_desig, "quantite": "123,5"}
        p = requests.patch(f"{BASE_URL}/api/dataset/{upload_id}/recap-row/{idx}", json=payload, timeout=30)
        assert p.status_code == 200 and p.json()["row"]["kind"] == "manual"

        # Export recap sheet
        r = requests.get(f"{BASE_URL}/api/export/{upload_id}?sheet=recap", timeout=120)
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
        assert "Récapitulatif" in wb.sheetnames
        ws = wb["Récapitulatif"]

        # Find row matching marker
        found = None
        for row in ws.iter_rows(values_only=True):
            if row and len(row) >= 4 and row[2] == marker_desig:
                found = row
                break
        wb.close()
        assert found is not None, "Manual row not found in exported Excel"
        # Columns: Type, Référence, Désignation, Quantité
        assert found[0] == "MTYPE"
        assert found[1] == "MREF99"
        assert found[2] == marker_desig
        assert found[3] == 123.5

        # cleanup
        requests.delete(f"{BASE_URL}/api/dataset/{upload_id}/recap-row/{idx}", timeout=30)
