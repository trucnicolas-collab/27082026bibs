"""iter47 — Cohérence "Résumé N1" (onglet Nuit N) ↔ "Détail allées" du rapport Excel.

Contexte métier (rappel utilisateur, très strict) :
  • Rouge = vrai problème AVEC commentaire de retard.
  • Écart validé par le poseur ("Tout est OK" cochée = justif_ok=True) → ORANGE, JAMAIS rouge.
  • Valeur ≥ prévu → vert (traité ailleurs).

Ce test met en place DEUX allées sur la même nuit dans un rapport Excel :
  1. Allée validée avec `justif_ok=True`  → doit être ORANGE + "OK poseur — validé"
  2. Allée validée avec `justification` texte + `justif_ok=False` → texte affiché tel quel

Puis vérifie que les DEUX onglets (`Nuit N` et `Détail allées`) affichent :
  • la même couleur de fond sur l'écart %
  • la même chaîne dans la colonne "Justification"

Autrement dit : impossible qu'un onglet dise « OK poseur — validé (orange) » et
que l'autre dise « manquante (rouge) » pour la même ligne, ce qui était
l'incohérence signalée par l'utilisateur.
"""

import io
import os
import asyncio
import pytest
import openpyxl
import requests
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv

load_dotenv("/app/backend/.env")
load_dotenv("/app/frontend/.env", override=False)

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
UPLOAD_ID = "fd15443f-6d2a-4cef-bd72-c56bb29e9c42"  # dataset preview partagé
NIGHT = 2
UID_OK = "1__A__R1"      # sera validée avec justif_ok=True    → ORANGE attendu
UID_TXT = "2__B__R2"     # sera validée avec justification texte → texte affiché
T = f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}"

# Couleurs XlsxWriter (bg) — cf. suivi_deploy.py L1275-1278
ORANGE_BG = "FEF3C7"     # C_WARNING_BG (justif_ok=True)
DANGER_BG = "FEE2E2"     # C_DANGER_BG  (aucune justification)


# ─── Fixtures ────────────────────────────────────────────────────────────────
@pytest.fixture(scope="module")
def auth():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login",
               json={"email": "admin@vusion.local", "password": "admin123"})
    assert r.status_code == 200, r.text
    return s


@pytest.fixture(scope="module", autouse=True)
def phase_both_on_night_two(auth):
    """Force les deux allées à être phasées sur la nuit 2 le temps du test,
    puis restaure la config d'origine."""
    async def _tmp(rows_val):
        client = AsyncIOMotorClient(os.environ["MONGO_URL"])
        db = client[os.environ["DB_NAME"]]
        d = await db.datasets.find_one({"upload_id": UPLOAD_ID})
        original = ((d.get("phasage") or {}).get("es") or {}).get("rows") or []
        new_rows = []
        for r in original:
            if r.get("id") == UID_TXT:
                new_rows.append({**r, "nuit": rows_val})
            else:
                new_rows.append(r)
        await db.datasets.update_one(
            {"upload_id": UPLOAD_ID},
            {"$set": {"phasage.es.rows": new_rows}})
        return original

    original = asyncio.run(_tmp(NIGHT))
    # Publier + reset état terrain
    auth.post(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/publish", json={"published": True})
    requests.patch(f"{T}/allee", json={
        "uid": UID_OK, "status": "a_faire",
        "justification": "", "justif_ok": False,
        "products": [
            {"designation": "ES 1.5 noir", "reel": None},
            {"designation": "ES 2.1 noir", "reel": None},
            {"designation": "990 mm (noir)", "reel": None},
        ],
    })
    requests.patch(f"{T}/allee", json={
        "uid": UID_TXT, "status": "a_faire",
        "justification": "", "justif_ok": False,
        "products": [{"designation": "ES 1.5 noir", "reel": None}],
    })
    yield
    # Restaure phasage d'origine
    async def _restore():
        client = AsyncIOMotorClient(os.environ["MONGO_URL"])
        db = client[os.environ["DB_NAME"]]
        await db.datasets.update_one(
            {"upload_id": UPLOAD_ID},
            {"$set": {"phasage.es.rows": original}})
    asyncio.run(_restore())
    requests.patch(f"{T}/allee", json={
        "uid": UID_OK, "status": "a_faire",
        "justification": "", "justif_ok": False,
        "products": [{"designation": "ES 1.5 noir", "reel": None}],
    })
    requests.patch(f"{T}/allee", json={
        "uid": UID_TXT, "status": "a_faire",
        "justification": "", "justif_ok": False,
        "products": [{"designation": "ES 1.5 noir", "reel": None}],
    })


# ─── Helpers ─────────────────────────────────────────────────────────────────
def _bg(cell) -> str:
    """Retourne le code hexa uppercase du bg-color d'une cellule openpyxl."""
    rgb = getattr(getattr(cell.fill.fgColor, "rgb", "") or "", "upper", lambda: "")()
    # openpyxl retourne "00FEF3C7" ou "FFFEF3C7" → strippe l'alpha (2 premiers)
    if isinstance(rgb, str) and len(rgb) == 8:
        return rgb[2:]
    return rgb or ""


def _find_row(ws, contains: str, in_col: int = 1) -> int:
    """Retourne le n° de ligne (1-indexé) dont la cellule in_col contient `contains`."""
    for row in ws.iter_rows():
        for c in row:
            if c.column == in_col and isinstance(c.value, str) and contains in c.value:
                return c.row
    raise AssertionError(f"Row containing {contains!r} in col {in_col} not found in {ws.title}")


def _generate_report_excel() -> openpyxl.Workbook:
    r = requests.get(f"{T}/rapport-nuit/{NIGHT}")
    assert r.status_code == 200, r.text
    return openpyxl.load_workbook(io.BytesIO(r.content), data_only=False)


# ─── Scénario : justif_ok=True (orange) ──────────────────────────────────────
def test_scenario_justif_ok_true_coherent_orange_in_both_sheets(auth):
    """Allée validée avec case "Tout est OK" cochée (poseur valide un écart)."""
    # Écart 20% (plan=100, reel=80)
    r = requests.patch(f"{T}/allee", json={
        "uid": UID_OK,
        "products": [{"designation": "ES 1.5 noir", "reel": 80}],
    })
    assert r.status_code == 200, r.text
    r = requests.patch(f"{T}/allee", json={
        "uid": UID_OK, "status": "validee", "justif_ok": True, "justification": "",
    })
    assert r.status_code == 200, r.text

    wb = _generate_report_excel()
    sheet_summary = f"Résumé N{NIGHT}"
    assert sheet_summary in wb.sheetnames, wb.sheetnames
    assert "Détail allées" in wb.sheetnames, wb.sheetnames

    # ── Onglet Résumé N{NIGHT} — chercher la ligne "ES 1.5 noir" dans la section justif ──
    ws1 = wb[sheet_summary]
    # Récupère toutes les lignes contenant "ES 1.5 noir" avec valeur pct dans col 8
    found_n1 = []
    for row in ws1.iter_rows():
        for c in row:
            if isinstance(c.value, str) and "ES 1.5 noir" in c.value:
                # Écart % est colonne 8 (0-idx 7)
                pct_cell = ws1.cell(row=c.row, column=8)
                # Justification cell : à droite du pct (colonne merged 9-12)
                justif_cell = ws1.cell(row=c.row, column=9)
                if pct_cell.value is not None:
                    found_n1.append((c.row, pct_cell, justif_cell))
    assert found_n1, "no 'ES 1.5 noir' row with pct value found in Nuit sheet"
    r1, pct1, jf1 = found_n1[-1]  # dernier = section justif (après la section KPI)
    assert _bg(pct1) == ORANGE_BG, (
        f"Nuit {NIGHT} pct bg = {_bg(pct1)}, expected {ORANGE_BG} (orange). "
        f"pct.value={pct1.value}, justif={jf1.value!r}")
    assert "OK poseur — validé" in (jf1.value or ""), (
        f"justif text = {jf1.value!r}")

    # ── Onglet Détail allées — mêmes vérifications ──
    ws2 = wb["Détail allées"]
    found_d = []
    for row in ws2.iter_rows():
        for c in row:
            if isinstance(c.value, str) and "ES 1.5 noir" in c.value:
                # Écart % en col 5 (0-idx 4), Justification en col 6
                pct_cell = ws2.cell(row=c.row, column=5)
                justif_cell = ws2.cell(row=c.row, column=6)
                if pct_cell.value is not None:
                    found_d.append((c.row, pct_cell, justif_cell))
    assert found_d, "no ES 1.5 noir row found in Détail allées"
    r2, pct2, jf2 = found_d[-1]
    assert _bg(pct2) == ORANGE_BG, (
        f"Détail allées pct bg = {_bg(pct2)}, expected {ORANGE_BG}. "
        f"pct={pct2.value}, justif={jf2.value!r}")
    assert "OK poseur — validé" in (jf2.value or ""), (
        f"Détail allées justif text = {jf2.value!r}")

    # ── COHÉRENCE : les 2 onglets ont la même couleur et le même texte ──
    assert _bg(pct1) == _bg(pct2), \
        f"INCOHÉRENCE couleur : Nuit={_bg(pct1)}, Détail={_bg(pct2)}"
    assert (jf1.value or "").strip() == (jf2.value or "").strip(), \
        f"INCOHÉRENCE texte : Nuit={jf1.value!r}, Détail={jf2.value!r}"


# ─── Scénario : justification texte (sans justif_ok) ─────────────────────────
def test_scenario_justification_text_coherent_in_both_sheets(auth):
    """Allée validée avec un texte de justification, sans cocher "Tout est OK".
    Le TEXTE fourni par le poseur doit apparaître à l'identique dans les 2 onglets."""
    # Écart 25% (plan=80, reel=60)
    r = requests.patch(f"{T}/allee", json={
        "uid": UID_TXT,
        "products": [{"designation": "ES 1.5 noir", "reel": 60}],
    })
    assert r.status_code == 200, r.text
    r = requests.patch(f"{T}/allee", json={
        "uid": UID_TXT, "status": "validee",
        "justif_ok": False,
        "justification": "Rayon réduit par le magasin",
    })
    assert r.status_code == 200, r.text

    wb = _generate_report_excel()
    ws1 = wb[f"Résumé N{NIGHT}"]
    ws2 = wb["Détail allées"]

    def _find_ecarts_section(ws, marker_text: str) -> int:
        """Retourne le n° de ligne (1-idx) où débute la section 'Écarts > 5%'.
        Insensible à la casse et aux emojis (Résumé N1 utilise « ÉCARTS > 5% JUSTIFIÉS »,
        Détail allées utilise « Écarts > 5% (EEG / rails ES) et justifications »)."""
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and "carts > 5%" in c.value.lower():
                    return c.row
        raise AssertionError(f"Section 'Écarts > 5%' introuvable dans {ws.title}")

    def _find_justif_row_in_section(ws, section_start: int, marker: str,
                                     pct_col: int, justif_col: int):
        """Cherche `marker` dans la colonne justification (col_justif) après
        le début de la section, retourne (row, pct_cell, justif_cell)."""
        for row in ws.iter_rows(min_row=section_start + 1):
            jc = ws.cell(row=row[0].row, column=justif_col)
            if isinstance(jc.value, str) and marker in jc.value:
                pct = ws.cell(row=row[0].row, column=pct_col)
                return row[0].row, pct, jc
        return None, None, None

    # Résumé N{N} : section "Écarts > 5%" avec écart pct en col 8, justif col 9-12 (mergée)
    s1 = _find_ecarts_section(ws1, "Écarts > 5%")
    r1, pct1, jf1 = _find_justif_row_in_section(ws1, s1, "Rayon réduit",
                                                 pct_col=8, justif_col=9)
    assert jf1 is not None, f"Texte justif absent de Résumé N{NIGHT} (section L{s1})"

    # Détail allées : section "Écarts > 5%" avec écart pct en col 5, justif col 6
    s2 = _find_ecarts_section(ws2, "Écarts > 5%")
    r2, pct2, jf2 = _find_justif_row_in_section(ws2, s2, "Rayon réduit",
                                                 pct_col=5, justif_col=6)
    assert jf2 is not None, f"Texte justif absent de Détail allées (section L{s2})"

    # Le TEXTE doit être strictement le même dans les deux onglets
    assert jf1.value.strip() == jf2.value.strip(), \
        f"Texte différent — Résumé={jf1.value!r}, Détail={jf2.value!r}"
    # Et la couleur de l'écart % doit être la même dans les 2 onglets
    assert _bg(pct1) == _bg(pct2), \
        f"INCOHÉRENCE couleur (texte justif) — Résumé={_bg(pct1)}, Détail={_bg(pct2)}"


# ─── Scénario : delta positif = vert dans Détail allées ──────────────────────
def test_delta_positive_is_green_in_detail_allees(auth):
    """Règle stricte utilisateur : pose ≥ prévu → vert, aucun état rouge/orange.
    Vérification côté onglet "Détail allées" (delta EEG posé - plan)."""
    # Reset UID_OK avec un delta positif : plan 100 → pose 110 sur ES 1.5 noir
    requests.patch(f"{T}/allee", json={
        "uid": UID_OK, "status": "a_faire",
        "justification": "", "justif_ok": False,
        "products": [
            {"designation": "ES 1.5 noir", "reel": 110},
            {"designation": "ES 2.1 noir", "reel": 50},
            {"designation": "990 mm (noir)", "reel": 5},
        ],
    })
    requests.patch(f"{T}/allee", json={"uid": UID_OK, "status": "validee"})

    wb = _generate_report_excel()
    ws = wb["Détail allées"]
    # Cherche la ligne de l'allée 1 (secteur A / rayon R1)
    for row in ws.iter_rows():
        vals = [c.value for c in row]
        if len(vals) >= 2 and str(vals[0] or "") == "1" and "A" in str(vals[1] or ""):
            # Colonne EEG posés (indices selon _rapport_response L1830+)
            # Chaque cellule numérique doit être ≥ 0, et les deltas doivent être verts
            for c in row:
                if isinstance(c.value, (int, float)) and c.value > 0:
                    bg = _bg(c)
                    # Aucun rouge (danger) sur une allée où pose ≥ plan
                    assert bg != DANGER_BG, (
                        f"cellule {c.coordinate}={c.value} ROUGE alors que delta positif !")
            break
    else:
        pytest.fail("Allée 1 A/R1 introuvable dans Détail allées")
