"""iter48h — Commentaire attaché à une photo (upload + edit + Excel).

Vérifie :
  1. POST /allee-photo avec form `comment` → la photo créée porte le commentaire
  2. PATCH /photo/{id} met à jour le commentaire d'une photo existante
  3. Le state expose photos[].comment
  4. L'Excel rapport-nuit contient une cellule caption sous la photo avec 💬 + texte
"""
import io
import os
import openpyxl
import pytest
import requests
from PIL import Image
from dotenv import load_dotenv

load_dotenv("/app/backend/.env")
load_dotenv("/app/frontend/.env", override=False)

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
UPLOAD_ID = "fd15443f-6d2a-4cef-bd72-c56bb29e9c42"
UID = "1__A__R1"
T = f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}"


def _tiny_jpeg() -> bytes:
    buf = io.BytesIO()
    Image.new("RGB", (200, 150), (120, 160, 200)).save(buf, format="JPEG", quality=70)
    return buf.getvalue()


@pytest.fixture(scope="module")
def auth():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login",
               json={"email": "admin@vusion.local", "password": "admin123"})
    assert r.status_code == 200, r.text
    return s


@pytest.fixture(scope="module", autouse=True)
def publish_and_cleanup(auth):
    auth.post(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/publish", json={"published": True})
    yield
    # Cleanup : supprime toutes les photos ajoutées par ce test
    r = requests.get(T)
    for a in (r.json() or {}).get("allees", []) or []:
        if a.get("uid") == UID:
            for p in (a.get("photos") or []):
                requests.delete(f"{T}/photo/{p['id']}")


def _post_photo(comment: str | None = None) -> str:
    files = {"file": ("test.jpg", _tiny_jpeg(), "image/jpeg")}
    data = {"uid": UID}
    if comment is not None:
        data["comment"] = comment
    r = requests.post(f"{T}/allee-photo", data=data, files=files)
    assert r.status_code == 200, r.text
    return r.json()["photo"]["id"]


def _get_photo_from_state(pid: str) -> dict | None:
    r = requests.get(T)
    for a in r.json().get("allees", []):
        for p in a.get("photos", []):
            if p["id"] == pid:
                return p
    return None


def test_upload_photo_avec_commentaire(auth):
    pid = _post_photo("Étagère abîmée à revoir")
    p = _get_photo_from_state(pid)
    assert p is not None, "photo introuvable dans le state"
    assert p["comment"] == "Étagère abîmée à revoir", p


def test_upload_photo_sans_commentaire_puis_patch(auth):
    pid = _post_photo(None)  # pas de comment initial
    p = _get_photo_from_state(pid)
    assert p["comment"] == "", p
    # Édite le commentaire a posteriori
    r = requests.patch(f"{T}/photo/{pid}", json={"comment": "Ajouté après coup"})
    assert r.status_code == 200, r.text
    p2 = _get_photo_from_state(pid)
    assert p2["comment"] == "Ajouté après coup", p2


def test_patch_commentaire_photo_inexistante():
    r = requests.patch(f"{T}/photo/nonexistent", json={"comment": "test"})
    assert r.status_code == 404


def test_commentaire_longueur_max_500(auth):
    """Le commentaire est tronqué à 500 caractères côté backend."""
    long = "x" * 800
    pid = _post_photo(long)
    p = _get_photo_from_state(pid)
    assert len(p["comment"]) == 500, len(p["comment"])


def test_excel_rapport_nuit_contient_caption_photo(auth):
    """Vérifie que l'Excel de nuit rend le commentaire sous la miniature."""
    pid = _post_photo("Commentaire visible dans Excel")
    r = requests.get(f"{T}/rapport-nuit/2")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=False)
    # Le Résumé N2 doit contenir une cellule contenant le texte du commentaire
    found = False
    for sname in [f"Résumé N2"]:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and "Commentaire visible dans Excel" in c.value:
                    # Vérifie le préfixe 💬 (indique une caption photo)
                    assert "💬" in c.value, c.value
                    found = True
                    break
            if found:
                break
    assert found, "caption photo introuvable dans Résumé N2"
