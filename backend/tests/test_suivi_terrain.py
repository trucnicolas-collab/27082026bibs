"""Backend tests for the new terrain (public upload_id) endpoints of /api/suivi + /api/suivi-terrain.

Refonte: le token PAR MAGASIN est SUPPRIMÉ. Remplacé par un espace commun basé
sur la publication (published=true) et clé upload_id.
"""
import io
import os
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL").rstrip("/")
UPLOAD_ID = "fd15443f-6d2a-4cef-bd72-c56bb29e9c42"
UID = "1__A__R1"


@pytest.fixture(scope="module")
def auth():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login",
               json={"email": "admin@vusion.local", "password": "admin123"})
    assert r.status_code == 200, r.text
    return s


@pytest.fixture(scope="module", autouse=True)
def ensure_published(auth):
    """Le magasin doit être publié pour ces tests, ET republié en fin de module."""
    r = auth.post(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/publish", json={"published": True})
    assert r.status_code == 200, r.text
    yield
    # cleanup: laisser publié à la fin
    auth.post(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/publish", json={"published": True})


# ---- publish + terrain-stores ------------------------------------------
def test_terrain_stores_lists_published(ensure_published):
    r = requests.get(f"{BASE_URL}/api/suivi-terrain/stores")
    assert r.status_code == 200
    j = r.json()
    assert "stores" in j
    ids = [s["upload_id"] for s in j["stores"]]
    assert UPLOAD_ID in ids, f"expected {UPLOAD_ID} in published stores; got {ids}"
    s = next(s for s in j["stores"] if s["upload_id"] == UPLOAD_ID)
    for k in ("upload_id", "store_name", "label", "filename", "published_by"):
        assert k in s
    assert s["published_by"] == "admin@vusion.local"


def test_publish_toggle_hides_and_shows(auth):
    # dépublier
    r = auth.post(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/publish", json={"published": False})
    assert r.status_code == 200
    assert r.json()["published"] is False
    # /stores ne doit plus le contenir
    ids = [s["upload_id"] for s in requests.get(f"{BASE_URL}/api/suivi-terrain/stores").json()["stores"]]
    assert UPLOAD_ID not in ids
    # terrain state → 404
    r = requests.get(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}")
    assert r.status_code == 404
    # republier
    r = auth.post(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/publish", json={"published": True})
    assert r.status_code == 200
    assert r.json()["published"] is True
    r = requests.get(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}")
    assert r.status_code == 200


def test_publish_requires_auth():
    r = requests.post(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/publish", json={"published": True})
    assert r.status_code in (401, 403)


# ---- publication metadata in chef state, absent from terrain state ----
def test_chef_state_has_publication(auth):
    j = auth.get(f"{BASE_URL}/api/suivi/{UPLOAD_ID}").json()
    assert "publication" in j
    assert j["publication"]["published"] is True
    assert j["publication"]["published_by"]


def test_terrain_state_hides_publication_key():
    j = requests.get(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}").json()
    assert j["is_terrain"] is True
    assert "publication" not in j
    assert "terrain" not in j
    assert "allees" in j
    assert j.get("geo_keys") == ["rails_es", "sa_15", "sa_21_std", "sa_21_freezer"]


def test_terrain_state_invalid_upload_id_404():
    r = requests.get(f"{BASE_URL}/api/suivi-terrain/deadbeef-dead-beef-dead-beefdeadbeef")
    assert r.status_code == 404


# ---- PATCH allee geoloc + comment (par PRODUIT) ------------------------
def test_terrain_patch_geo_gap_and_alert_explanation():
    r = requests.patch(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}/allee",
                       json={"uid": UID, "geoloc_comment": "",
                             "products": [{"designation": "990 mm (noir)", "reel": 5, "geo": 3}]})
    assert r.status_code == 200, r.text

    j = requests.get(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}").json()
    a = next(a for a in j["allees"] if a["uid"] == UID)
    assert a["reel"]["rails_es"] == 5
    assert a["geo"]["rails_es"] == 3
    assert a["geo_gap"].get("rails_es") == 2

    alerts_geo = [al for al in j["alerts"]
                  if al.get("type") == "geoloc" and al.get("uid") == UID
                  and al.get("family") in (None, "rails_es")]
    assert alerts_geo
    assert alerts_geo[0]["needs_explanation"] is True

    # ajouter commentaire
    r = requests.patch(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}/allee",
                       json={"uid": UID, "geoloc_comment": "TEST_ explication automatique"})
    assert r.status_code == 200
    j = requests.get(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}").json()
    alerts_geo = [al for al in j["alerts"]
                  if al.get("type") == "geoloc" and al.get("uid") == UID
                  and al.get("family") in (None, "rails_es")]
    assert alerts_geo[0]["needs_explanation"] is False


def test_terrain_patch_negative_geo_422():
    r = requests.patch(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}/allee",
                       json={"uid": UID,
                             "products": [{"designation": "990 mm (noir)", "geo": -1}]})
    assert r.status_code == 422


# ---- Photos public ------------------------------------------------------
def test_terrain_photo_upload_get_delete():
    img_bytes = _tiny_png()
    r = requests.post(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}/allee-photo",
                      files={"file": ("test.png", img_bytes, "image/png")},
                      data={"uid": UID})
    assert r.status_code == 200, r.text
    pid = r.json()["photo"]["id"]

    r = requests.get(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}/photo/{pid}")
    assert r.status_code == 200
    assert r.headers.get("content-type", "").startswith("image/")
    assert len(r.content) > 0

    j = requests.get(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}").json()
    a = next(a for a in j["allees"] if a["uid"] == UID)
    assert any(p["id"] == pid for p in a["photos"])

    r = requests.post(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}/allee-photo",
                      files={"file": ("x.txt", b"hello", "text/plain")},
                      data={"uid": UID})
    assert r.status_code == 400

    r = requests.delete(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}/photo/{pid}")
    assert r.status_code == 200


def test_auth_photo_endpoints(auth):
    r = auth.post(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/allee-photo",
                  files={"file": ("test.png", _tiny_png(), "image/png")},
                  data={"uid": UID})
    assert r.status_code == 200
    pid = r.json()["photo"]["id"]
    r = auth.get(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/photo/{pid}")
    assert r.status_code == 200 and len(r.content) > 0
    r = auth.delete(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/photo/{pid}")
    assert r.status_code == 200


# ---- Rapport terrain ----------------------------------------------------
def test_terrain_rapport_nuit_has_geo_and_images():
    r = requests.post(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}/allee-photo",
                      files={"file": ("test.png", _tiny_png(), "image/png")},
                      data={"uid": UID})
    assert r.status_code == 200
    pid = r.json()["photo"]["id"]
    try:
        r = requests.get(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}/rapport-nuit/2")
        assert r.status_code == 200
        assert r.content[:2] == b"PK"
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        flat = "\n".join(" | ".join(str(c) if c is not None else "" for c in row)
                        for row in ws.iter_rows(values_only=True))
        assert "Géoloc" in flat
        assert "Commentaire GÉOLOC" in flat
        assert len(ws._images) >= 1
    finally:
        requests.delete(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}/photo/{pid}")


# ---- Incidents ---------------------------------------------------------
def test_terrain_incident_create_delete():
    r = requests.post(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}/incident",
                      json={"nuit": 2, "text": "TEST_terrain incident"})
    assert r.status_code == 200
    inc_id = r.json()["incident"]["id"]
    j = requests.get(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}").json()
    assert any(i["id"] == inc_id for i in j["incidents"])
    r = requests.delete(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}/incident/{inc_id}")
    assert r.status_code == 200


# ---- Reset (créateur/admin only) ---------------------------------------
def test_reset_requires_auth():
    r = requests.delete(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/reset")
    assert r.status_code in (401, 403)


def test_reset_by_non_creator_forbidden():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login",
               json={"email": "poseur@test.local", "password": "poseur123"})
    assert r.status_code == 200
    r = s.delete(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/reset")
    assert r.status_code in (403, 404), r.text


def test_reset_by_creator_clears_data_keeps_publication(auth):
    # create some data first
    r = auth.patch(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/allee",
                   json={"uid": UID,
                         "products": [{"designation": "990 mm (noir)", "reel": 3}]})
    assert r.status_code == 200
    r = auth.post(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/incident",
                  json={"nuit": 2, "text": "TEST_before_reset"})
    assert r.status_code == 200

    # reset
    r = auth.delete(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/reset")
    assert r.status_code == 200, r.text

    # data cleared
    j = auth.get(f"{BASE_URL}/api/suivi/{UPLOAD_ID}").json()
    # progression should be zero: sum reel == 0 across allees
    total_reel = 0
    for a in j.get("allees", []):
        for v in (a.get("reel") or {}).values():
            total_reel += (v or 0)
    assert total_reel == 0
    assert len(j.get("incidents") or []) == 0

    # publication kept
    assert j["publication"]["published"] is True


# ---- terrain endpoints on unpublished return 404 -----------------------
def test_terrain_endpoints_404_when_unpublished(auth):
    # unpublish
    auth.post(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/publish", json={"published": False})
    try:
        r = requests.patch(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}/allee",
                           json={"uid": UID,
                                 "products": [{"designation": "990 mm (noir)", "reel": 1}]})
        assert r.status_code == 404
        r = requests.post(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}/incident",
                          json={"nuit": 2, "text": "x"})
        assert r.status_code == 404
        r = requests.get(f"{BASE_URL}/api/suivi-terrain/{UPLOAD_ID}/materiel")
        assert r.status_code == 404
    finally:
        auth.post(f"{BASE_URL}/api/suivi/{UPLOAD_ID}/publish", json={"published": True})


# ---------------------------------------------------------------- helpers
def _tiny_png():
    return (b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
            b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\xcf\xc0"
            b"\x00\x00\x00\x03\x00\x01[\x9f\x8c\x0c\x00\x00\x00\x00IEND\xaeB`\x82")
