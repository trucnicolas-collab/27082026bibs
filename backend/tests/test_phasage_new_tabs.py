"""Tests for new phasage tabs: Phasage caméras / Phasage full / Suivi phasage.

Covers:
- POST /api/upload-excel
- GET  /api/dataset/{id}/phasage-summary  (nested {es,cam,suivi})
- PATCH /api/dataset/{id}/phasage         (nested payload, persistence)
- GET  /api/export/{id}?sheet=all|phasage_cam|phasage_full|suivi
- Excel re-opens with openpyxl
- No banned formulas (TEXTJOIN, dynamic arrays, etc.)
"""

import io
import os
import re
import pytest
import requests
from openpyxl import load_workbook

def _load_backend_url():
    url = os.environ.get("REACT_APP_BACKEND_URL")
    if not url:
        # Fallback: read from frontend/.env (test environment)
        env_path = "/app/frontend/.env"
        if os.path.exists(env_path):
            with open(env_path) as fh:
                for line in fh:
                    if line.startswith("REACT_APP_BACKEND_URL="):
                        url = line.split("=", 1)[1].strip()
                        break
    if not url:
        raise RuntimeError("REACT_APP_BACKEND_URL is not set")
    return url.rstrip("/")


BASE_URL = _load_backend_url()
SAMPLE_PATH = "/tmp/sample_full.xlsx"

BANNED_FORMULAS = re.compile(r"\b(TEXTJOIN|FILTER|UNIQUE|XLOOKUP|TEXTSPLIT|SORT|SORTBY|SEQUENCE|LET|LAMBDA)\b", re.I)


@pytest.fixture(scope="module")
def api():
    s = requests.Session()
    return s


@pytest.fixture(scope="module")
def upload_id(api):
    with open(SAMPLE_PATH, "rb") as f:
        r = api.post(f"{BASE_URL}/api/upload-excel",
                     files={"file": ("sample_full.xlsx", f,
                                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    data = r.json()
    assert "upload_id" in data or "id" in data
    return data.get("upload_id") or data.get("id")


# --- 1. Upload + summary ------------------------------------------------------
class TestPhasageSummary:
    def test_summary_has_nested_phasage(self, api, upload_id):
        r = api.get(f"{BASE_URL}/api/dataset/{upload_id}/phasage-summary")
        assert r.status_code == 200
        d = r.json()
        assert "phasage" in d
        ph = d["phasage"]
        assert set(ph.keys()) >= {"es", "cam", "suivi"}
        assert "nb_nuits" in ph["es"] and "rows" in ph["es"]
        assert "nb_nuits" in ph["cam"] and "rows" in ph["cam"]
        assert "start_at_nuit" in ph["cam"]
        assert "rows" in ph["suivi"]

    def test_summary_totals_cameras_positive(self, api, upload_id):
        r = api.get(f"{BASE_URL}/api/dataset/{upload_id}/phasage-summary")
        d = r.json()
        assert d["totals"].get("cameras", 0) > 0, f"Expected cameras > 0, got {d['totals']}"


# --- 2. PATCH phasage nested + persistence -----------------------------------
class TestPhasagePatch:
    def test_patch_nested_persists(self, api, upload_id):
        # get current
        cur = api.get(f"{BASE_URL}/api/dataset/{upload_id}/phasage-summary").json()
        allees = [a["allee"] for a in cur["allees"]]
        cam_allees = [a["allee"] for a in cur["allees"] if (a.get("cameras") or 0) > 0]
        assert len(allees) >= 1
        assert len(cam_allees) >= 1

        payload = {
            "es": {"nb_nuits": 4, "rows": [
                {"id": "es-1", "allee": allees[0], "nuit": 1},
                {"id": "es-2", "allee": allees[-1], "nuit": 2},
            ]},
            "cam": {"nb_nuits": 3, "start_at_nuit": 5, "rows": [
                {"id": "cam-1", "allee": cam_allees[0], "nuit": 1},
            ]},
            "suivi": {"rows": [
                {"nuit": 1, "es_reel": 10.5, "cam_reel": None, "rails_geoloc": 4},
                {"nuit": 5, "es_reel": None, "cam_reel": 250, "rails_geoloc": None},
            ]},
        }
        r = api.patch(f"{BASE_URL}/api/dataset/{upload_id}/phasage", json=payload)
        assert r.status_code == 200, r.text
        echo = r.json()["phasage"]
        assert echo["es"]["nb_nuits"] == 4
        assert echo["cam"]["start_at_nuit"] == 5
        assert echo["cam"]["nb_nuits"] == 3
        assert len(echo["suivi"]["rows"]) == 2

        # GET to verify persistence
        again = api.get(f"{BASE_URL}/api/dataset/{upload_id}/phasage-summary").json()["phasage"]
        assert again["es"]["nb_nuits"] == 4
        assert again["cam"]["start_at_nuit"] == 5
        assert any(r["allee"] == cam_allees[0] and r["nuit"] == 1
                   for r in again["cam"]["rows"])
        suivi_n1 = next((r for r in again["suivi"]["rows"] if r["nuit"] == 1), None)
        assert suivi_n1 is not None
        assert suivi_n1["es_reel"] == 10.5
        assert suivi_n1["rails_geoloc"] == 4
        suivi_n5 = next((r for r in again["suivi"]["rows"] if r["nuit"] == 5), None)
        assert suivi_n5 is not None and suivi_n5["cam_reel"] == 250


# --- 3. Exports --------------------------------------------------------------
def _check_no_banned_formulas(wb):
    bad = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    if BANNED_FORMULAS.search(v):
                        bad.append((sn, cell.coordinate, v))
    return bad


class TestExport:
    def test_export_all_contains_sheets(self, api, upload_id):
        r = api.get(f"{BASE_URL}/api/export/{upload_id}", params={"sheet": "all"})
        assert r.status_code == 200
        assert r.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        wb = load_workbook(io.BytesIO(r.content), data_only=False)
        names = set(wb.sheetnames)
        for required in ("Phasage de pose", "Phasage caméras", "Phasage full", "Suivi phasage"):
            assert required in names, f"Missing sheet {required!r}. Got: {names}"

    def test_export_all_no_banned_formulas(self, api, upload_id):
        r = api.get(f"{BASE_URL}/api/export/{upload_id}", params={"sheet": "all"})
        wb = load_workbook(io.BytesIO(r.content), data_only=False)
        bad = _check_no_banned_formulas(wb)
        assert not bad, f"Found banned formulas: {bad[:5]}"

    @pytest.mark.parametrize("sheet,expected", [
        ("phasage_cam", "Phasage caméras"),
        ("phasage_full", "Phasage full"),
        ("suivi", "Suivi phasage"),
    ])
    def test_export_single_sheet(self, api, upload_id, sheet, expected):
        r = api.get(f"{BASE_URL}/api/export/{upload_id}", params={"sheet": sheet})
        assert r.status_code == 200, r.text[:300]
        wb = load_workbook(io.BytesIO(r.content), data_only=False)
        assert expected in wb.sheetnames

    def test_export_phasage_cam_formulas(self, api, upload_id):
        r = api.get(f"{BASE_URL}/api/export/{upload_id}", params={"sheet": "phasage_cam"})
        wb = load_workbook(io.BytesIO(r.content), data_only=False)
        ws = wb["Phasage caméras"]
        formulas = [c.value for row in ws.iter_rows() for c in row
                    if isinstance(c.value, str) and c.value.startswith("=")]
        assert any("VLOOKUP" in f for f in formulas), "Expected VLOOKUP formulas"
        assert any("SUMIFS" in f for f in formulas), "Expected SUMIFS formulas"
        assert all(not BANNED_FORMULAS.search(f) for f in formulas)

    def test_export_suivi_diff_formulas(self, api, upload_id):
        r = api.get(f"{BASE_URL}/api/export/{upload_id}", params={"sheet": "suivi"})
        wb = load_workbook(io.BytesIO(r.content), data_only=False)
        ws = wb["Suivi phasage"]
        formulas = [c.value for row in ws.iter_rows() for c in row
                    if isinstance(c.value, str) and c.value.startswith("=")]
        # Diff formulas
        assert any("IFERROR" in f and "-" in f for f in formulas), formulas[:5]
        assert all(not BANNED_FORMULAS.search(f) for f in formulas)
